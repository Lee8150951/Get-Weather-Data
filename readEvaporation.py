from netCDF4 import Dataset
from typing import List
import xlrd
import openpyxl
from numpy import *
import os

# 封装阅读参数函数，参数列表：1.文件路径；2.传入经度；3.传入纬度
def readData(path: str, longitude: float, latitude: float) -> float:
    nc_obj = Dataset(path)
    # 解析数据源
    lonList = list(nc_obj.variables['longitude'][:])
    latList = list(nc_obj.variables['latitude'][:])
    e = nc_obj.variables['e'][:][0]
    # 查询当前经纬度信息
    p_lon = lonList.index(longitude)
    p_lat = latList.index(latitude)
    # 获取数据
    data = e[p_lat][p_lon]
    return data

# 封装站点信息查询函数
def siteData() -> List[int]:
    data = xlrd.open_workbook('./export/position/long&lati.xls')
    table = data.sheets()[0]
    row = table.nrows
    # 按行读取
    for i in range(row):
        # 获取行数据
        rowdate = table.row_values(i)
        # 处理行数据
        while '' in rowdate:
            rowdate.remove('')
    return row

# 封装删除多余sheet函数，参数列表：1.文件路径
# 解决多创建一个sheet的问题
def delSheet(path: str):
    delWb = openpyxl.load_workbook(path)
    ws = delWb["Sheet"]
    delWb.remove(ws)
    delWb.save(path)

# 数据处理函数封装
def factory(source: float) -> float:
    # 更换单位mm
    source *= 1000
    source = '%.5f' % source
    return source
    

# 封装分城市创建excel文件函数，参数列表：1.城市名称；2.数据列表
def cityDate(city: str, data: List[int]):
    newExl = openpyxl.Workbook()
    sheet = newExl.create_sheet()
    # 创建表头
    months = ["年份\月份", "一月", "二月", "三月", "四月", "五月", "六月", "七月", "八月", "九月", "十月", "十一月", "十二月"]
    sheet.append(months)
    # 循环数据（含表头，每13个一个大循环）
    for i in range(0, 273, 13):
        # 每一年的数据
        yearData = data[i : i + 13]
        # 将数据存入excel中
        sheet.append(yearData)
    # 储存
    path = f"./export/evaporation/{city}.xlsx"
    newExl.save(filename = path)

if __name__ == '__main__':
    # 循环文件
    staticName_1 = "G:\\妮的数据\\气象数据\\C602\\"
    staticName_2 = "monthly-era5-single-levels-"
    months = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
    
    # 循环站点信息
    tableData = xlrd.open_workbook('./export/position/long&lati.xls')
    table = tableData.sheets()[0]
    row = table.nrows
    # 按行读取
    for i in range(row):
        # 获取行数据
        rowdata = table.row_values(i)
        # 处理行数据
        while '' in rowdata:
            rowdata.remove('')
        # 获取地级市名
        city = rowdata[0]
        # Log输出
        logStr = f"当前正在处理的城市是：{city}..."
        print(logStr)
        # 地级市数据仓库
        storage = list()
        # 循环文件
        for year in range(2000, 2021):
            storage.append(year)
            for month in months:
                # 临时存储
                tempo = list()
                # 循环站点
                for current in range(1, len(rowdata)):
                    # 获取并拆分经纬度
                    lola = rowdata[current].split(",")
                    # 经度；纬度
                    longitude, latitude = float(lola[0]), float(lola[1])
                    path = staticName_1 + str(year) + "\\" + staticName_2 + str(year) + month + ".nc"
                    # 获取当前文件的当前站点信息
                    data = readData(path, longitude, latitude)
                    tempo.append(data)
                # 求平均值
                ave = mean(tempo)
                # 数据处理
                factoryData = factory(ave)
                storage.append(factoryData)
        cityDate(city, storage)
    print("数据处理完成！")
    os.system("pause")