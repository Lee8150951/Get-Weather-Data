import xlrd
import openpyxl

# 处理地级市经纬度
data = xlrd.open_workbook('./resources/positions.xls')
# 获取表格数据(从sheet中)
table = data.sheets()[0]
# 临时仓库
storage = list()
# 获取数据
positions = table.col_values(0)[1:]
for position in positions:
    current = position.split(",")
    current[2] = current[2][1:-1]
    storage.append(current)
    
# 将数据导入新excel中
newExl = openpyxl.Workbook()
sheet = newExl.create_sheet()
# 表头
row0 = ['地点', '经度', '纬度']
sheet.append(row0)
# 添加数据
for i in storage:
    i_position, i_longitude, i_latitude = i[2], i[0], i[1]
    i_current = [i_position, i_longitude, i_latitude]
    sheet.append(i_current)
# 储存
newExl.save(filename = './export/position/test.xlsx')

# 删除多余sheet
delWb = openpyxl.load_workbook('./export/position/test.xlsx')
ws = delWb["Sheet"]
delWb.remove(ws)
delWb.save('./export/position/test.xlsx')

print("Success!")